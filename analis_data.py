from openpyxl import load_workbook
import csv
import pandas as pd

# загрузка партнеров
spisok_azs = pd.read_excel('список АЗС.xlsx')
spisok_azs = spisok_azs.to_numpy()
for i in spisok_azs:
    i[2] = i[2].split(',')[0]
    i[0] = str(i[0])

def partner_AZS(num_azs):
    for i in spisok_azs:
        if num_azs == i[0]:
            return i[1]

def regions(azs_n):
    for i in spisok_azs:
        if azs_n in i:
            return i[2]


data = [] # для итогового датафрейма
data_for_ds = [] # для дипсика
data_for_tonalnost = [] # для дипсика


#===================================================================================
#============================== ЗАГРУЗКА НОВЫХ ДАННЫХ ==============================
#===================================================================================


# Файлы для обработки
#file = 'rocketData.xlsx' # название файла с выгрузкой
file = 'rd.xlsx'
workbook = load_workbook(file)
sheets = workbook.sheetnames

sheet_otz = workbook[sheets[0]]

for row in sheet_otz.iter_rows(min_row=4, values_only=True):
    id = row[12]
    address = str(row[7]) + ', ' + str(row[8])
    date = row[0]
    number_azs = row[2]
    region = regions(number_azs)
    source = row[9]
    rank = row[10]
    text = row[13]
    time_answer = row[16]
    answer = row[17]
    responsible = row[18]
    partner = partner_AZS(number_azs)
    data.append( [ id, date, number_azs, partner, region, address, source, rank, text, time_answer, answer, responsible ] )
    data_for_tonalnost.append([date, number_azs, rank, text])
    data_for_ds.append([date, number_azs, rank, text])

new_data = pd.DataFrame(data, columns=['id Отзыва', 'Дата', "АЗС", "Партнер", "Регион", "Адрес", "Каталог", "Рейтинг", "Отзыв", "Время обработки отзыва", "Ответ", "Ответственный"])
old_data = pd.read_excel('ДатаФрейм.xlsx')
all = pd.concat([old_data, new_data], ignore_index=False)

all.to_excel('ДатаФрейм.xlsx', index=False)


#===================================================================================
#============================== ПРИСВОЕНИЕ ТЕГОВ ===================================
#===================================================================================

from openai import OpenAI

client = OpenAI(api_key="sk-2caddd3f28cb4226b4a4798886c2d53d", base_url="https://api.deepseek.com") # СЕКРЕТНЫЙ КЛЮЧ

for i in data_for_ds:
    if type(i[3]) == str:
        prompt = f"""Тебе необходимо определить на что жалуется клиент или что хвалит клиент в отзыве. Темы, которые могут быть затронуты в комментарии:  
        Хамство персонала, Организация продаж наливных нефтепродуктов, Действия оператора, Недолив, Поверхностный комментарий, Плохой туалет(негатив), Оплата, Благодарность топливо, Проблема с бонусами, Благодарность общая, Отсутствие кассира, Отсутствие автозаправщиков, Антисанитария, Негатив общий, Цена, Перелив, Негатив по внешним данным азс, Плохое качество топлива, Негатив персонал, Благодарность персоналу. Определи темы комментария, который будет приведен ниже, свои темы не придумывай. На один комментарий может быть несколько тем. Рассуждай по ходу. В ответе предоставь только список тем, где каждая написана с новой строки. Комментарий \n{i[3]}"""
        response = client.chat.completions.create(
            model="deepseek-chat",
            messages=[
                {"role": "system", "content": "Ты аналитик в крупной компании."},
                {"role": "user", "content": prompt},
            ],
            stream=False,
            temperature=1
        )
        print("Ответ модели:", response.choices[0].message.content)
        tegi = response.choices[0].message.content.split('\n')
        for t in tegi:
            i.append(t)
    if data_for_ds.index(i) % 100 == 0:
        df = pd.DataFrame(data_for_ds)
        df.to_excel('ОбработаноНЕЙРОпроцесс.xlsx', index=False)

df = pd.DataFrame(data_for_ds)
df.to_excel('ОбработаноНЕЙРОпроцесс.xlsx', index=False)

#===================================================================================
#========================== ОПРЕДЕЛЕНИЕ ТОНАЛЬНОСТИ ================================
#===================================================================================


for i in data_for_tonalnost:
    if type(i[3]) == str:
        prompt = f"""Тебе необходимо определить тональность отзыва. Тональности, которые могут быть: негативная, нейтральная, позитивная. Свои тональности не придумывай. В ответ предоставь только тональность комментария одним словом (негативная/нейтральная/позитивная). 
Отзыв: \n{i[3]}"""
        response = client.chat.completions.create(
            model="deepseek-chat",
            messages=[
                {"role": "system", "content": "Ты аналитик в крупной компании."},
                {"role": "user", "content": prompt},
            ],
            stream=False,
            temperature=1
        )
        print("Ответ модели:", response.choices[0].message.content)
        i.append(response.choices[0].message.content)

    if data_for_tonalnost.index(i) % 100 == 0:
        df = pd.DataFrame(data_for_tonalnost)
        df.to_excel('ТональностьНЕЙРОпроцесс.xlsx', index=False)

df = pd.DataFrame(data_for_tonalnost)
df.to_excel('ТональностьНЕЙРОпроцесс.xlsx', index=False)

#===================================================================================
#======================= СОЗДАНИЕ ТАБЛИЦ ДЛЯ ДАШБОРДОВ =============================
#===================================================================================

# Создание отдельных таблиц
data_rocket = data[:201] # УБРАТЬ ПОСЛЕ ДЕМОНСТРАЦИИ
for i in range(len(data_rocket)):
    data_rocket[i] = [data_rocket[i][0], data_rocket[i][1], data_rocket[i][2], data_rocket[i][7], data_rocket[i][8]]

all_teg = pd.read_excel('ОбработаноНЕЙРОпроцесс.xlsx')
all_teg

ALL_teg = all_teg.to_numpy()
a = pd.DataFrame(ALL_teg)
ALL_teg = a.to_numpy()
ALL_teg[-1]
d = len(ALL_teg[0])

com_and_teg = []
for i in data_rocket:
    for j in ALL_teg:
        try:
            if i[1] == j[0] and str(i[2]) == str(j[1]) and int(i[3]) == int(j[2]):
                for t in range(d, 0):
                    if type(j[t]) == str and j[t] != j[0] and j[t] != j[1] and j[t] != j[2]:
                        com_and_teg.append([i[0], j[t].strip()])
        except:
            if i[1] == j[0] and str(i[2]) == str(j[1]) and i[4] == j[3]:
                for t in range(d, 0):
                    if type(j[t]) == str and j[t] != j[0] and j[t] != j[1] and j[t] != j[2]:
                        com_and_teg.append([i[0], j[t].strip()])


kategorii = pd.read_excel('Категории.xlsx')


category_dict = {column: kategorii[column].dropna().tolist() for column in kategorii.columns}


def choice_category(teg):
    for category in category_dict:
        for t in category_dict[category]:
            if teg.lower() == t.lower():
                return category

for i in com_and_teg:
    cat = choice_category(i[1])
    i.append(cat)

table_com_and_teg = pd.DataFrame(com_and_teg, columns=['id Отзыва', "Тег", "Категория"])
old_table = pd.read_excel('Комментарии_теги.xlsx')
all = pd.concat([old_table, table_com_and_teg], ignore_index=False)

#all.to_excel('Комментарии_тегиПРОВЕРКА.xlsx', index=False, header=['id Отзыва', "Тег", "Категория"])
all.to_excel('Комментарии_теги.xlsx', index=False, header=['id Отзыва', "Тег", "Категория"])



all_ton = pd.read_excel('ТональностьНЕЙРОпроцесс.xlsx')


list_all_ton = all_ton.to_numpy()

com_and_ton = []
for i in data_rocket:
    for j in list_all_ton:
        try:
            if i[1] == j[0] and str(i[2]) == str(j[1]) and int(i[3]) == int(j[2]):
                com_and_ton.append([i[0], j[4]])
        except:
            if i[1] == j[0] and str(i[2]) == str(j[1]) and i[4] == j[3]:
                com_and_ton.append([i[0], j[4]])

table_com_and_ton = pd.DataFrame(com_and_ton, columns=['id Отзыва', "Тональность"])
old_table = pd.read_excel('Комментарии_тональность.xlsx')
all = pd.concat([old_table, table_com_and_ton], ignore_index=False)

#all.to_excel("Комментарии_тональностьПРОВЕРКА.xlsx", index=False, header=['id Отзыва', "Тональность"])
all.to_excel("Комментарии_тональность.xlsx", index=False, header=['id Отзыва', "Тональность"])
